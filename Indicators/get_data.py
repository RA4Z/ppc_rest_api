import json
import openpyxl
import datetime
import Indicators.format_stocks
import Indicators.format_ovs
import Indicators.format_efficiency

target_year = int((datetime.date.today() - datetime.timedelta(days=datetime.date.today().day)).strftime("%Y"))


class Excel:
    def __init__(self, file: str, sheet: str):
        self.workbook = openpyxl.load_workbook(file, data_only=True, keep_vba=True)
        self.sheet = self.workbook[sheet]

    def count_rows(self, column: int):
        max_row = self.sheet.max_row
        for row in range(max_row, 0, -1):
            cell_value = self.sheet.cell(row, column).value
            if cell_value is not None:
                return row
        return 0

    def get_cell(self, row: int, column: int):
        return self.sheet.cell(row, column).value

    def count_columns(self, row: int):
        max_col = self.sheet.max_column
        for col in range(max_col, 0, -1):
            cell_value = self.sheet.cell(row=row, column=col).value
            if cell_value is not None:
                return col
        return 0


def atualizar_stocks():
    arquivo = ("\\\\intranet.weg.net@SSL\\DavWWWRoot\\br\\energia-wm\\pcp\\Comisso de Estoques\\Gráficos de giro de "
               f"estoque\\InventoryManagement_ENERGIA {target_year} .xlsm")
    excel = Excel(arquivo, 'Database')
    rows = excel.count_rows(1)
    columns = excel.count_columns(1)
    header = [cell.value for cell in excel.sheet[1]]  # Primeira linha como cabeçalho

    data = []

    for index in range(2, rows):
        row_data = {}
        for col in range(1, columns + 1):
            row_data[header[col - 1]] = excel.get_cell(index, col)
        data.append(row_data)

    data = Indicators.format_stocks.formatar_json(data)

    with open("Indicators/data/stocks.json", "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)


def atualizar_atendimento_ov():
    arquivo = (
        f"\\\\intranet.weg.net@SSL\\DavWWWRoot\\br\\energia-wm\\pcp\\PQWP/PWQP - {target_year}\\Depto Planejamento e "
        f"Controle da Produção {target_year}.xlsx")
    excel = Excel(arquivo, 'Planilha1')
    rows = excel.count_rows(2)
    columns = excel.count_columns(2)
    header = [cell.value for cell in excel.sheet[2]]  # Primeira linha como cabeçalho

    data = []

    for index in range(2, rows):
        row_data = {}
        for col in range(2, columns + 1):
            row_data[header[col - 1]] = excel.get_cell(index, col)
        data.append(row_data)

    data = Indicators.format_ovs.formatar_json(data)

    with open("Indicators/data/OVs.json", "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)


def atualizar_efficiency():
    arquivo = ("Q:\\GROUPS\\BR_SC_JGS_WM_LOGISTICA\\PCP\\Suely\\Indicadores Diretoria Industrial\\INDICADORES "
               f"{target_year}\\Global ID Efficiency {target_year}.xlsm")
    excel = Excel(arquivo, 'Plan1')
    rows = excel.count_rows(1)
    columns = excel.count_columns(4)
    header = [cell.value for cell in excel.sheet[4]]

    data = []

    for index in range(5, rows + 1):
        row_data = {}
        for col in range(1, columns + 1):
            row_data[header[col - 1]] = excel.get_cell(index, col)
        data.append(row_data)

    data = Indicators.format_efficiency.formatar_json(data)

    with open("Indicators/data/efficiency.json", "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)


def mesclar_dados():
    data = (json.load(open("Indicators/data/stocks.json", "r", encoding="utf-8"))
            + json.load(open("Indicators/data/OVs.json", "r", encoding="utf-8"))
            + json.load(open("Indicators/data/efficiency.json", "r", encoding="utf-8")))

    with open("Indicators/rest/wen_indicators_database.json", "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)


def encontrar_empresa_por_titulo(title):
    data = json.load(open("Indicators/data/indicadores.json", "r", encoding="utf-8"))
    for item in data:
        if title in item.get('indicadores', []):
            return item.get('empresa'), item.get('id')
    return None, None


def resumir_info():
    data = json.load(open("Indicators/rest/wen_indicators_database.json", "r", encoding="utf-8"))
    indicadores = ['On time Delivery', 'Inventory*', 'Inventory Turns', 'Efficiency']
    resultado = []
    for info in data:
        title = info['Concatenar'] if info.get('Concatenar') else info['Indicador']
        ultimo_valor = info['averages'][-1]

        for ind in indicadores:
            if ind in title:
                empresa, country = encontrar_empresa_por_titulo(title)  # Passar 'data' para a função
                result = ultimo_valor / info['Meta'] if info.get('Meta') else None
                resultado.append({'company': empresa, 'indicator': ind, 'title': title, 'average': ultimo_valor,
                                  'target': info['Meta'], 'result': result, 'country': country})
                break  # Se encontrou o indicador, passa para o próximo título

    with open("Indicators/rest/wen_indicators_results.json", "w", encoding="utf-8") as json_file:
        json.dump(resultado, json_file, ensure_ascii=False, indent=4)


if __name__ == "__main__":
    resumir_info()
